#!/usr/bin/env python3
"""
fill_template.py — Preenche um template XLSX com dados de um CSV.

Sintaxe das células no template: {nome_coluna}
O script localiza a célula com {nome_coluna}, substitui pela primeira linha de dados
e preenche as células abaixo com as linhas subsequentes.

Uso: fill_template.py <template.xlsx> <output.xlsx> <dados.csv>
"""
import sys
import re
import openpyxl
import pandas as pd


def fill_template(template_path, output_path, csv_path):
    df = pd.read_csv(csv_path, sep=';', encoding='utf-8-sig', dtype=str).fillna('')
    wb = openpyxl.load_workbook(template_path)

    for ws in wb.worksheets:
        marker_row = None
        markers = {}  # col_index (1-based) -> df_column_name

        for row_cells in ws.iter_rows():
            for cell in row_cells:
                if not (cell.value and isinstance(cell.value, str)):
                    continue
                m = re.match(r'^\{(\w+)\}$', cell.value.strip())
                if not m:
                    continue
                col_name = m.group(1)
                if col_name not in df.columns:
                    continue
                if marker_row is None:
                    marker_row = cell.row
                markers[cell.column] = col_name
                cell.value = None  # limpa o marcador

        if not markers:
            continue

        # Preenche dados a partir de marker_row
        for i, (_, df_row) in enumerate(df.iterrows()):
            target_row = marker_row + i
            for col_idx, col_name in markers.items():
                ws.cell(row=target_row, column=col_idx, value=df_row[col_name])

    wb.save(output_path)


if __name__ == '__main__':
    if len(sys.argv) != 4:
        print(f'Uso: {sys.argv[0]} <template.xlsx> <output.xlsx> <dados.csv>', file=sys.stderr)
        sys.exit(1)
    fill_template(sys.argv[1], sys.argv[2], sys.argv[3])
    print('OK')
